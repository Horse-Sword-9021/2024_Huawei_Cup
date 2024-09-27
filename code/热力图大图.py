import numpy as np
import openpyxl
import seaborn as sns
import matplotlib.pyplot as plt
from scipy.stats import binned_statistic_2d
from scipy.ndimage import gaussian_filter
import matplotlib as mpl


def pic_gener(start_stamp, end_stamp, file_name, sheet_name, plot_index):
    workbook = openpyxl.load_workbook(file_name)
    worksheet = workbook[sheet_name]

    total = [[worksheet.cell(start_stamp,2).value, worksheet.cell(start_stamp,1).value, 1]]
    # 截取时间片段

    for row in worksheet.iter_rows(min_row=start_stamp+1, max_row=end_stamp, values_only=True):
        point1 = [row[1], row[0]]
        count = 0
        for point in total:
            if point1 == point[:2]:
                point[2] += 1
                count = 1
                break
            else:
                continue
        if count == 0:
            total.append([row[1], row[0], 1])

    x = []
    y = []
    weight = []
    for item in total:
        x.append(item[0])
        y.append(item[1])
        weight.append(item[2])
    max_index = np.argsort(weight)[::-1]

    max_center = [x[max_index[0]], y[max_index[0]]]
    max_speed = max_center[1]*9/max_center[0]

    weight = np.array(weight)
    mean_weight = np.mean(weight)
    std_weight = np.std(weight)
    # Z-score 归一化
    normalized_weight = (weight - mean_weight) / std_weight

    xmax = np.max(x)
    xmin = np.min(x)
    ymax = np.max(y)
    ymin = np.min(y)

    # 计算二维直方图
    x_bins = np.linspace(xmin, xmax, 20)  # 增加 bins 数量
    y_bins = np.linspace(ymin, ymax, 20)
    stat = binned_statistic_2d(x, y, normalized_weight, 'sum', bins=[x_bins, y_bins])

    # 对统计数据应用高斯模糊
    stat_smooth = gaussian_filter(stat[0], sigma=1)

    # 绘制平滑后的热力图
    axs[plot_index].imshow(stat_smooth.T, extent=(xmin, xmax, ymin, ymax), origin='lower', cmap='viridis', aspect='auto')

    mpl.rcParams['font.sans-serif'] = ['SimHei']
    mpl.rcParams['axes.unicode_minus'] = False
    axs[plot_index].set_xlim(xmin, xmax)
    axs[plot_index].set_ylim(ymin, ymax)
    axs[plot_index].grid(False)

    # 绘制60km线和40km线
    # 设置 x 值
    xline = np.linspace(xmin, xmax, 100)

    # 计算 y 值
    yline1 = 6.67 * xline  # k = 6.67 的直线
    yline2 = 4.44 * xline  # k = 4.44 的直线
    yline3 = 2.22 * xline  # k = 2.22 的直线

    return max_center, max_speed


if __name__ == '__main__':
    file_name1 = '31_1135_1352.xlsx'
    sheet_name1 = 'Sheet2'
    # pic_gener(2, 45, file_name1, sheet_name1)
    # 以15为一个时间戳，即现实中的5min
    workbook = openpyxl.load_workbook(file_name1)
    worksheet = workbook[sheet_name1]
    max_row = worksheet.max_row - 1
    workbook.close()
    cut = 15
    initial = 2
    stamp_cut = []
    while initial < max_row:
        stamp_cut.append([initial, initial + cut])
        initial += cut
    stamp_cut[-1][-1] = max_row
    # 生成大图
    # 参数设置
    num_plots = len(stamp_cut)  # 假设我们要生成12个图
    cols = 6  # 每行的列数
    rows = (num_plots + cols - 1) // cols  # 计算需要的行数
    # 创建一个大图和子图
    fig, axs = plt.subplots(rows, cols, figsize=(15, 10))
    axs = axs.flatten()  # 将二维数组展平成一维数组
    i = 0
    max_center_group = []
    for single_cut in stamp_cut:
        c, s = pic_gener(single_cut[0], single_cut[1], file_name1, sheet_name1, i)
        max_center_group.append([c[0], c[1], s])
        i += 1
    # 隐藏最后两个子图
    for i in range(34, 36):
        axs.flatten()[i].axis('off')
        # 调整布局

    plt.savefig('热力大图不加线', dpi = 800)
    pass
